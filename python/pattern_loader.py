# -*- coding: utf-8 -*-
"""
안테나 패턴 로더.

기본 입력은 ``Data/pattern/260827_pattern.xlsx``의 ``야기+옴니`` 시트다.
이 시트에는 사용자가 구성한 이중 야기 + 옴니 합성 패턴이 0..360도로
정리되어 있다. 이전 OM900 CSV도 명시적인 ``path`` 인자로 계속 읽을 수 있다.

CSV 구조
--------
deg, 910MHz 수평, 910MHz 수직, 955MHz 수평, 955MHz 수직

여기서 '수평/수직'은 편파가 아니라 측정 plane(cut)으로 해석한다.

- 수평: Horizontal-plane / Azimuth pattern
- 수직: Vertical-plane / Elevation pattern

첨부 데이터의 형태도 이 해석과 일치한다.
수평면은 비교적 omnidirectional이고,
수직면은 theta ~= 90도 부근에서 최대이고 천정/저면에 null이 있다.
"""

from __future__ import annotations

import codecs
import csv
import os
from typing import Dict, List

import numpy as np


DEFAULT_PATTERN_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "..", "Data", "pattern", "260827_pattern.xlsx",
)
DEFAULT_PATTERN_SHEET = "야기+옴니"
DEFAULT_PATTERN_MODEL = "이중 야기 + 옴니 (260827)"
PATTERN_MODELS = {
    "야기안테나": "야기 안테나 (260827)",
    "옴니": "옴니 안테나 (260827)",
    "야기+옴니": DEFAULT_PATTERN_MODEL,
}

FREQ_BANDS = {
    "910": {"horizontal_col": 1, "vertical_col": 2},
    "955": {"horizontal_col": 3, "vertical_col": 4},
}


def _read_rows(path: str) -> List[List[str]]:
    """cp949 우선, 그 다음 UTF 계열로 CSV를 읽는다."""
    last_error = None
    for enc in ("cp949", "utf-8-sig", "utf-8"):
        try:
            with codecs.open(path, "r", enc) as f:
                return list(csv.reader(f))
        except UnicodeDecodeError as e:
            last_error = e

    raise IOError(f"패턴 CSV 인코딩을 판별할 수 없습니다: {path}") from last_error


def _read_xlsx_pattern(path: str, sheet_name: str = DEFAULT_PATTERN_SHEET):
    """Excel 합성 시트의 angle/수평/수직 열을 읽는다."""
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise RuntimeError(
            "260827 패턴 Excel을 읽으려면 openpyxl이 필요합니다."
        ) from error

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"패턴 시트를 찾을 수 없습니다: {sheet_name} "
                f"(가능: {', '.join(workbook.sheetnames)})"
            )
        sheet = workbook[sheet_name]
        entries = {}
        for angle, horizontal, vertical in sheet.iter_rows(
            min_row=2,
            max_col=3,
            values_only=True,
        ):
            try:
                entries[float(angle)] = (float(horizontal), float(vertical))
            except (TypeError, ValueError):
                continue
        return PATTERN_MODELS.get(sheet_name, sheet_name), entries
    finally:
        workbook.close()


def _read_csv_pattern(path: str, freq: str):
    rows = _read_rows(path)
    cols = FREQ_BANDS[freq]
    model = ""
    if rows:
        if len(rows[0]) > 1 and rows[0][1].strip():
            model = rows[0][1].strip()
        if len(rows) > 1 and rows[1][0].strip():
            model = rows[1][0].strip()

    entries = {}
    for row in rows[3:]:
        try:
            deg = float(row[0])
            g_h = float(row[cols["horizontal_col"]])
            g_v = float(row[cols["vertical_col"]])
        except (ValueError, IndexError):
            continue
        entries[deg] = (g_h, g_v)
    return model, entries


def _power_average_db(values_db):
    """dB 값들을 linear power에서 평균한 뒤 다시 dB로 변환."""
    values_db = np.asarray(values_db, dtype=float)
    p = np.power(10.0, values_db / 10.0)
    return float(10.0 * np.log10(np.mean(p)))


def load_pattern(
    path: str | None = None,
    freq: str = "955",
    sheet_name: str = DEFAULT_PATTERN_SHEET,
) -> Dict:
    if path is None:
        path = DEFAULT_PATTERN_PATH

    freq = str(freq)
    if freq not in FREQ_BANDS:
        raise ValueError(
            f"지원하지 않는 주파수 대역: {freq} "
            f"(가능: {', '.join(FREQ_BANDS)})"
        )

    absolute_path = os.path.abspath(path)
    is_xlsx = os.path.splitext(absolute_path)[1].lower() in (".xlsx", ".xlsm")
    if is_xlsx:
        model, entries = _read_xlsx_pattern(absolute_path, sheet_name=sheet_name)
    else:
        model, entries = _read_csv_pattern(absolute_path, freq)

    if not entries:
        raise ValueError(f"유효한 패턴 데이터가 없습니다: {path}")

    raw_angles = np.asarray(sorted(entries.keys()), dtype=float)
    horizontal_db = np.asarray([entries[a][0] for a in raw_angles], dtype=float)
    vertical_db = np.asarray([entries[a][1] for a in raw_angles], dtype=float)

    # ------------------------------------------------------------
    # 수직면 패턴을 Sionna theta(0..180 deg)로 변환
    #
    # 원본 vertical cut은 -180..179 deg의 한 바퀴 데이터다.
    # 3D 근사에서는 azimuth 의존성을 horizontal cut이 담당하므로,
    # vertical cut의 +theta / -theta를 같은 elevation으로 접어(fold)
    # linear-power 평균한다.
    # ------------------------------------------------------------
    v_by_angle = {float(a): float(entries[a][1]) for a in entries}

    theta_deg = np.arange(0.0, 181.0, 1.0)
    theta_gain_db = []

    for theta in theta_deg:
        candidates = []

        if is_xlsx:
            # 0..360 polar cut: theta와 360-theta는 같은 고도각의 반대쪽 cut이다.
            if theta in v_by_angle:
                candidates.append(v_by_angle[theta])
            mirror = 360.0 - theta
            if mirror in v_by_angle and mirror != theta:
                candidates.append(v_by_angle[mirror])
        else:
            if theta in v_by_angle:
                candidates.append(v_by_angle[theta])
            neg = -theta
            if neg in v_by_angle and neg != theta:
                candidates.append(v_by_angle[neg])
            # CSV에는 +180 대신 -180만 있는 경우가 일반적
            if theta == 180.0 and -180.0 in v_by_angle:
                candidates.append(v_by_angle[-180.0])

        if not candidates:
            raise ValueError(f"수직면 패턴에서 theta={theta}° 값을 만들 수 없습니다.")

        theta_gain_db.append(_power_average_db(candidates))

    theta_gain_db = np.asarray(theta_gain_db, dtype=float)

    # ------------------------------------------------------------
    # 수평면 패턴
    # 원본 -180..179 deg를 azimuth(phi)로 그대로 사용.
    # 보간 시 seam을 닫기 위해 +180점은 -180점과 동일하게 추가한다.
    # ------------------------------------------------------------
    if is_xlsx:
        # 0..359를 Sionna/브라우저 공통 방위축 -180..180으로 변환한다.
        phi_pairs = []
        for angle, gains in entries.items():
            if angle >= 360.0:
                continue
            phi = ((float(angle) + 180.0) % 360.0) - 180.0
            phi_pairs.append((phi, float(gains[0])))
        phi_pairs.sort(key=lambda item: item[0])
        phi_deg = np.asarray([item[0] for item in phi_pairs], dtype=float)
        phi_gain_db = np.asarray([item[1] for item in phi_pairs], dtype=float)
        if len(phi_deg) and phi_deg[0] == -180.0:
            phi_deg = np.append(phi_deg, 180.0)
            phi_gain_db = np.append(phi_gain_db, phi_gain_db[0])
    else:
        phi_deg = raw_angles.copy()
        phi_gain_db = horizontal_db.copy()
        if -180.0 in entries and 180.0 not in entries:
            phi_deg = np.append(phi_deg, 180.0)
            phi_gain_db = np.append(phi_gain_db, entries[-180.0][0])

    max_h = float(np.max(phi_gain_db))
    max_v = float(np.max(theta_gain_db))

    # 2D cut들을 합성할 때 기준이 되는 절대 peak gain.
    # 둘 중 큰 측정 최대값을 사용한다.
    target_peak = max(max_h, max_v)

    return {
        "model": model,
        "freq_mhz": int(freq),
        "configuration": "dual Yagi + omni",
        "source_file": os.path.basename(absolute_path),
        "source_sheet": sheet_name if is_xlsx else None,

        "raw_angle_deg": raw_angles.tolist(),
        "raw_horizontal_gain_dbi": horizontal_db.tolist(),
        "raw_vertical_gain_dbi": vertical_db.tolist(),

        "phi_deg": phi_deg.tolist(),
        "horizontal_gain_dbi": phi_gain_db.tolist(),

        "theta_deg": theta_deg.tolist(),
        "vertical_gain_dbi": theta_gain_db.tolist(),

        "horizontal_max_gain_dbi": max_h,
        "vertical_max_gain_dbi": max_v,
        "max_gain_dbi": target_peak,
    }


if __name__ == "__main__":
    p = load_pattern(freq="955")
    print("model:", p["model"])
    print("frequency:", p["freq_mhz"], "MHz")
    print("horizontal max:", round(p["horizontal_max_gain_dbi"], 2), "dBi")
    print("vertical max:", round(p["vertical_max_gain_dbi"], 2), "dBi")
    print("3D target peak:", round(p["max_gain_dbi"], 2), "dBi")
    print("theta range:", p["theta_deg"][0], "~", p["theta_deg"][-1])
    print("phi range:", p["phi_deg"][0], "~", p["phi_deg"][-1])
